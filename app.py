import re
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from datetime import datetime
from zoneinfo import ZoneInfo

import math

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components


# =========================
# CONFIGURACIÓN
# =========================
COL_RESPUESTA_F = 6
COL_RESPUESTA_G = 7
COL_RESPUESTA_K = 11

VALID_RESPUESTAS_F = {"SI", "DESARROLLO", "TERCERO", "NO"}
VALID_RESPUESTAS_G = {"SI", "NO"}
VALID_RESPUESTAS_K = {"COMPLETO", "PARCIAL", "INCOMPLETO"}

HOJAS_EXCLUIDAS = {"1.15", "1.16", "1.17"}


# =========================
# FUNCIONES
# =========================
def normalizar(valor):
    if valor is None:
        return ""
    return str(valor).strip().upper()


def es_hoja_1x(nombre):
    nombre = nombre.strip()
    prefijo = re.match(r"^(\d+\.\d+)", nombre)
    if prefijo and prefijo.group(1) in HOJAS_EXCLUIDAS:
        return False
    return bool(re.match(r"^1\.", nombre))


def es_hoja_no_funcional(nombre):
    nombre = nombre.strip()
    prefijo = re.match(r"^(\d+\.\d+)", nombre)
    if prefijo and prefijo.group(1) in HOJAS_EXCLUIDAS:
        return True
    return False


def detectar_filas(ws):
    filas = []
    for r in range(9, ws.max_row + 1):
        val = normalizar(ws.cell(r, 1).value)
        if "*** FIN DEL DOCUMENTO ***" in val:
            break
        if val.isdigit():
            filas.append(r)
    return filas


def leer_respuesta(ws, fila, col, validas):
    val = ws.cell(fila, col).value
    if val is None:
        return "VACIO"
    val = normalizar(val)
    if val in validas:
        return val
    return "VACIO"


def analizar_hoja(ws, pesos_f, peso_col_f, peso_col_g):
    data = []
    for r in detectar_filas(ws):
        resp_f = leer_respuesta(ws, r, COL_RESPUESTA_F, VALID_RESPUESTAS_F)
        resp_g = leer_respuesta(ws, r, COL_RESPUESTA_G, VALID_RESPUESTAS_G)
        requerimiento = ws.cell(r, 4).value
        peso_f = pesos_f.get(resp_f, 0.0)
        peso_g = peso_col_g if resp_g == "SI" else 0.0
        data.append({
            "Hoja": ws.title,
            "Fila": r,
            "Requerimiento": requerimiento,
            "Peso_F": peso_f,
            "Peso_G": peso_g,
            "Peso_Total": peso_f + peso_g
        })
    df = pd.DataFrame(data)
    if df.empty:
        return None, None
    maximo_posible = peso_col_f + peso_col_g
    if maximo_posible == 0:
        return 0.0, df
    cumplimiento = (df["Peso_Total"].mean() / maximo_posible) * 100
    return round(cumplimiento, 2), df


def analizar_hoja_k(ws, pesos_k):
    data = []
    for r in detectar_filas(ws):
        resp_k = leer_respuesta(ws, r, COL_RESPUESTA_K, VALID_RESPUESTAS_K)
        requerimiento = ws.cell(r, 4).value
        peso_k = pesos_k.get(resp_k, 0.0)
        data.append({
            "Hoja": ws.title,
            "Fila": r,
            "Requerimiento": requerimiento,
            "Peso_K": peso_k
        })
    df = pd.DataFrame(data)
    if df.empty:
        return None, None
    maximo_posible = pesos_k.get("COMPLETO", 1.0)
    if maximo_posible == 0:
        return 0.0, df
    calidad = (df["Peso_K"].mean() / maximo_posible) * 100
    return round(calidad, 2), df


def analizar_archivo(path, pesos_f, peso_col_f, peso_col_g, pesos_k, incluir_calidad):
    wb = openpyxl.load_workbook(path, data_only=True)

    hojas_func = [s for s in wb.sheetnames if es_hoja_1x(s)]
    resultados, detalles, resultados_k, detalles_k = {}, {}, {}, {}
    for h in hojas_func:
        ws = wb[h]
        cumplimiento, detalle_df = analizar_hoja(ws, pesos_f, peso_col_f, peso_col_g)
        if cumplimiento is not None:
            resultados[h] = cumplimiento
            detalles[h] = detalle_df
        if incluir_calidad:
            calidad, detalle_k_df = analizar_hoja_k(ws, pesos_k)
            if calidad is not None:
                resultados_k[h] = calidad
                detalles_k[h] = detalle_k_df

    hojas_nofunc = [s for s in wb.sheetnames if es_hoja_no_funcional(s)]
    resultados_nf, detalles_nf, resultados_k_nf, detalles_k_nf = {}, {}, {}, {}
    for h in hojas_nofunc:
        ws = wb[h]
        cumplimiento, detalle_df = analizar_hoja(ws, pesos_f, peso_col_f, peso_col_g)
        if cumplimiento is not None:
            resultados_nf[h] = cumplimiento
            detalles_nf[h] = detalle_df
        if incluir_calidad:
            calidad, detalle_k_df = analizar_hoja_k(ws, pesos_k)
            if calidad is not None:
                resultados_k_nf[h] = calidad
                detalles_k_nf[h] = detalle_k_df

    return (resultados, detalles, resultados_k, detalles_k,
            resultados_nf, detalles_nf, resultados_k_nf, detalles_k_nf)


def analizar_hoja_experiencia(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("3.")), None)
    if hoja_nombre is None:
        return pd.DataFrame([{"Sector/Industria": "", "País": "", "Proveedor": proveedor}])
    ws = wb[hoja_nombre]
    data = []
    for r in range(11, ws.max_row + 1):
        num = ws.cell(r, 2).value
        if num is None:
            continue
        sector = ws.cell(r, 4).value
        pais   = ws.cell(r, 5).value
        if sector is None and pais is None:
            continue
        data.append({
            "Sector/Industria": str(sector).strip() if sector else "",
            "País": str(pais).strip() if pais else "",
            "Proveedor": proveedor
        })
    if not data:
        return pd.DataFrame([{"Sector/Industria": "", "País": "", "Proveedor": proveedor}])
    return pd.DataFrame(data)


def analizar_hoja_experiencia_oferente(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("5.")), None)
    if hoja_nombre is None:
        return pd.DataFrame([{"Sector/Industria": "", "País": "", "Proveedor": proveedor}])
    ws = wb[hoja_nombre]
    data = []
    for r in range(11, ws.max_row + 1):
        num = ws.cell(r, 2).value
        if num is None:
            continue
        sector = ws.cell(r, 4).value
        pais   = ws.cell(r, 5).value
        if sector is None and pais is None:
            continue
        data.append({
            "Sector/Industria": str(sector).strip() if sector else "",
            "País": str(pais).strip() if pais else "",
            "Proveedor": proveedor
        })
    if not data:
        return pd.DataFrame([{"Sector/Industria": "", "País": "", "Proveedor": proveedor}])
    return pd.DataFrame(data)


def analizar_hoja_alcance_servicios(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("6.")), None)
    if hoja_nombre is None:
        return pd.DataFrame([{"Respuesta": "", "Proveedor": proveedor}])
    ws = wb[hoja_nombre]
    data = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 3).value
        if val is None:
            continue
        val_norm = str(val).strip().upper()
        if val_norm not in {"SI", "NO"}:
            continue
        nombre = ws.cell(r, 2).value
        if nombre is None:
            continue
        data.append({"Respuesta": val_norm, "Proveedor": proveedor})
    if not data:
        return pd.DataFrame([{"Respuesta": "", "Proveedor": proveedor}])
    return pd.DataFrame(data)


def analizar_hoja_metodologia(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("7.")), None)
    if hoja_nombre is None:
        return pd.DataFrame([{"Respuesta": "", "Proveedor": proveedor}])
    ws = wb[hoja_nombre]
    data = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 3).value
        if val is None:
            continue
        val_norm = str(val).strip().upper()
        if val_norm not in {"SI", "NO"}:
            continue
        nombre = ws.cell(r, 2).value
        if nombre is None:
            continue
        data.append({"Respuesta": val_norm, "Proveedor": proveedor})
    if not data:
        return pd.DataFrame([{"Respuesta": "", "Proveedor": proveedor}])
    return pd.DataFrame(data)


def construir_tablas(data, data_k, peso_total_cumplimiento, peso_total_calidad, incluir_calidad):
    df = pd.DataFrame(data)
    df_final = df.pivot_table(
        index="Hoja", columns="Proveedor", values="Cumplimiento_%", aggfunc="first"
    ).fillna(0).reset_index()
    total_por_proveedor = df.groupby("Proveedor")["Cumplimiento_%"].mean().round(2).to_dict()
    df_total = pd.DataFrame([{"Hoja": "TOTAL", **total_por_proveedor}]).fillna(0)

    df_final_k, df_total_k, df_puntaje, df_total_puntaje = None, None, None, None

    if incluir_calidad and data_k:
        df_k = pd.DataFrame(data_k)
        df_final_k = df_k.pivot_table(
            index="Hoja", columns="Proveedor", values="Calidad_%", aggfunc="first"
        ).fillna(0).reset_index()
        total_k_por_proveedor = df_k.groupby("Proveedor")["Calidad_%"].mean().round(2).to_dict()
        df_total_k = pd.DataFrame([{"Hoja": "TOTAL", **total_k_por_proveedor}]).fillna(0)

        proveedores = [c for c in df_final.columns if c != "Hoja"]
        df_puntaje = df_final[["Hoja"]].copy()
        df_final_k_idx = df_final_k.set_index("Hoja")
        df_final_idx = df_final.set_index("Hoja")
        for prov in proveedores:
            cum = df_final_idx[prov] if prov in df_final_idx.columns else 0
            cal = df_final_k_idx[prov] if prov in df_final_k_idx.columns else 0
            df_puntaje[prov] = ((cum * peso_total_cumplimiento) + (cal * peso_total_calidad)).round(2).values
        total_puntaje = {prov: round(df_puntaje[prov].mean(), 2) for prov in proveedores}
        df_total_puntaje = pd.DataFrame([{"Hoja": "TOTAL", **total_puntaje}])

    return df_final, df_total, df_final_k, df_total_k, df_puntaje, df_total_puntaje


def formatear_porcentaje_df(df):
    df_fmt = df.copy()
    for col in df_fmt.columns:
        if col != "Hoja":
            df_fmt[col] = df_fmt[col].apply(lambda x: f"{x:.2f}%")
    return df_fmt


def df_to_excel_bytes(dfs: dict) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def boton_descarga(label, dfs: dict, file_name: str, key: str):
    st.download_button(
        label,
        df_to_excel_bytes(dfs),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key
    )


def obtener_fecha_modificacion(archivo_bytes):
    import zipfile
    import xml.etree.ElementTree as ET
    import re as _re

    try:
        with zipfile.ZipFile(BytesIO(archivo_bytes)) as z:
            with z.open("docProps/core.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {"dcterms": "http://purl.org/dc/terms/"}
                modified_el = root.find("dcterms:modified", ns)
                if modified_el is None or not modified_el.text:
                    return "No disponible"

                raw = modified_el.text.strip()

                formatos = [
                    "%Y-%m-%dT%H:%M:%SZ",
                    "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%dT%H:%M",
                    "%Y-%m-%d",
                ]
                dt = None
                for fmt in formatos:
                    try:
                        dt = datetime.strptime(raw, fmt).replace(tzinfo=ZoneInfo("UTC"))
                        break
                    except ValueError:
                        continue

                if dt is None:
                    try:
                        clean = _re.sub(r"Z$", "+00:00", raw)
                        dt = datetime.fromisoformat(clean)
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
                    except ValueError:
                        pass

                if dt is not None:
                    dt_local = dt.astimezone(ZoneInfo("America/Bogota"))
                    return dt_local.strftime("%Y-%m-%d %H:%M:%S")

    except Exception:
        pass
    return "No disponible"


def construir_hoja_info_analisis(
    fecha_generacion,
    incluir_calidad,
    peso_col_f,
    peso_col_g,
    pesos_f,
    pesos_k,
    peso_total_cumplimiento,
    peso_total_calidad,
    pesos_hojas_func,
    pesos_hojas_nf,
    metadata_archivos,
    peso_alcance,       # ← NUEVO
    peso_metodologia,   # ← NUEVO
):
    bloques = []

    df_fecha = pd.DataFrame([{
        "Fecha y hora de generación del reporte": fecha_generacion
    }])
    bloques.append(("Generación del reporte", df_fecha))

    params_generales = [
        {"Parámetro": "Incluir calidad en el análisis", "Valor": "Sí" if incluir_calidad else "No"},
        {"Parámetro": "Peso máximo columna F (cubrimiento)", "Valor": peso_col_f},
        {"Parámetro": "Peso máximo columna G (inclusión)",   "Valor": peso_col_g},
    ]
    df_params = pd.DataFrame(params_generales)
    bloques.append(("Parámetros generales", df_params))

    pesos_f_rows = [
        {"Respuesta": k, "Peso aplicado": v}
        for k, v in pesos_f.items()
    ]
    df_pesos_f = pd.DataFrame(pesos_f_rows)
    bloques.append(("Pesos cubrimiento (columna F)", df_pesos_f))

    if incluir_calidad:
        pesos_k_rows = [
            {"Respuesta": k, "Peso aplicado": v}
            for k, v in pesos_k.items()
        ]
        df_pesos_k = pd.DataFrame(pesos_k_rows)
        bloques.append(("Pesos calidad (columna K)", df_pesos_k))

        df_pesos_totales = pd.DataFrame([
            {"Parámetro": "Peso total cumplimiento", "Valor": peso_total_cumplimiento},
            {"Parámetro": "Peso total calidad",      "Valor": peso_total_calidad},
        ])
        bloques.append(("Pesos totales puntaje combinado", df_pesos_totales))

    if pesos_hojas_func:
        df_ph_func = pd.DataFrame([
            {"Hoja": h, "Peso asignado (%)": p}
            for h, p in pesos_hojas_func.items()
        ])
        bloques.append(("Pesos por hoja funcional", df_ph_func))

    if pesos_hojas_nf:
        df_ph_nf = pd.DataFrame([
            {"Hoja": h, "Peso asignado (%)": p}
            for h, p in pesos_hojas_nf.items()
        ])
        bloques.append(("Pesos por hoja no funcional", df_ph_nf))

    # ── NUEVO: bloque alcance y metodología ──────────────────────────────────
    df_pesos_adicionales = pd.DataFrame([
        {"Parámetro": "Peso alcance",      "Valor": peso_alcance},
        {"Parámetro": "Peso metodología",  "Valor": peso_metodologia},
    ])
    bloques.append(("Pesos alcance y metodología", df_pesos_adicionales))
    # ─────────────────────────────────────────────────────────────────────────

    if metadata_archivos:
        df_archivos = pd.DataFrame(metadata_archivos)
        bloques.append(("Archivos analizados", df_archivos))

    return bloques


def escribir_hoja_info_analisis(writer, bloques):
    ws = writer.book.create_sheet("Informacion de analisis")
    fila_actual = 1

    for titulo, df in bloques:
        cell = ws.cell(row=fila_actual, column=1, value=titulo)
        cell.font = openpyxl.styles.Font(bold=True, size=11)
        fila_actual += 1

        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=fila_actual, column=col_idx, value=col_name).font = \
                openpyxl.styles.Font(bold=True)
        fila_actual += 1

        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=fila_actual, column=col_idx, value=value)
            fila_actual += 1

        fila_actual += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


# =========================
# UI
# =========================
st.set_page_config(layout="wide")
st.title("Analizador de Cumplimiento")

if st.sidebar.button("Reiniciar análisis"):
    st.session_state.clear()
    st.rerun()

with st.sidebar:
    if "incluir_calidad" not in st.session_state:
        st.session_state["incluir_calidad"] = False

    incluir_calidad = st.toggle(
        "Agregar calidad al análisis",
        value=st.session_state["incluir_calidad"],
        key="toggle_calidad"
    )
    st.session_state["incluir_calidad"] = incluir_calidad

    st.divider()

    st.header("Pesos cumplimiento funcional")
    st.caption("Todos los pesos se ingresan de 0 a 100 (se convierten internamente a escala 0.0–1.0)")

    peso_col_f_pct = st.number_input(
        "Peso máximo columna F (cubrimiento) — rango: 0 a 100",
        min_value=0, max_value=100, value=100, step=5,
        key="ni_peso_col_f"
    )
    peso_col_g_pct = st.number_input(
        "Peso máximo columna G (inclusión) — rango: 0 a 100",
        min_value=0, max_value=100, value=100, step=5,
        key="ni_peso_col_g"
    )

    peso_col_f = peso_col_f_pct / 100
    peso_col_g = peso_col_g_pct / 100

    st.markdown("**Pesos cubrimiento (proporción del peso máximo col F):**")
    st.caption("Cada valor indica qué porcentaje del peso máximo de col F se asigna a esa respuesta. Rango: 0 a 100")

    _si_pct = st.number_input(
        "SI (columna F) — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, value=100, step=5,
        key="ni_si_pct"
    )
    _desarrollo_pct = st.number_input(
        "DESARROLLO (columna F) — rango: 0 a 100, pred: 50",
        min_value=0, max_value=100, value=50, step=5,
        key="ni_desarrollo_pct"
    )
    _tercero_pct = st.number_input(
        "TERCERO (columna F) — rango: 0 a 100, pred: 50",
        min_value=0, max_value=100, value=50, step=5,
        key="ni_tercero_pct"
    )
    _no_pct = st.number_input(
        "NO (columna F) — rango: 0 a 100, pred: 0",
        min_value=0, max_value=100, value=0, step=5,
        key="ni_no_pct"
    )

    pesos_f = {
        "SI":         (_si_pct         / 100) * peso_col_f,
        "DESARROLLO": (_desarrollo_pct / 100) * peso_col_f,
        "TERCERO":    (_tercero_pct    / 100) * peso_col_f,
        "NO":         (_no_pct         / 100) * peso_col_f,
        "VACIO": 0.0
    }

    if incluir_calidad:
        st.divider()
        st.markdown("**Pesos calidad (columna K):**")
        st.caption("Rango: 0 a 100")

        _k_completo = st.number_input(
            "Completo (columna K) — rango: 0 a 100, pred: 100",
            min_value=0, max_value=100, value=100, step=5,
            key="ni_k_completo"
        )
        _k_parcial = st.number_input(
            "Parcial (columna K) — rango: 0 a 100, pred: 50",
            min_value=0, max_value=100, value=50, step=5,
            key="ni_k_parcial"
        )
        _k_incompleto = st.number_input(
            "Incompleto (columna K) — rango: 0 a 100, pred: 0",
            min_value=0, max_value=100, value=0, step=5,
            key="ni_k_incompleto"
        )

        pesos_k = {
            "COMPLETO":   _k_completo   / 100,
            "PARCIAL":    _k_parcial    / 100,
            "INCOMPLETO": _k_incompleto / 100,
            "VACIO": 0.0
        }

        st.markdown("**Pesos totales del puntaje combinado:**")
        st.caption("Rango: 0 a 100")

        _peso_total_cumplimiento_pct = st.number_input(
            "Peso total cumplimiento — rango: 0 a 100, pred: 100",
            min_value=0, max_value=100, value=100, step=5,
            key="ni_peso_total_cum"
        )
        _peso_total_calidad_pct = st.number_input(
            "Peso total calidad — rango: 0 a 100, pred: 100",
            min_value=0, max_value=100, value=100, step=5,
            key="ni_peso_total_cal"
        )

        peso_total_cumplimiento = _peso_total_cumplimiento_pct / 100
        peso_total_calidad      = _peso_total_calidad_pct      / 100

    else:
        pesos_k = {"COMPLETO": 1.0, "PARCIAL": 0.5, "INCOMPLETO": 0.0, "VACIO": 0.0}
        peso_total_cumplimiento = 1.0
        peso_total_calidad = 1.0

    st.divider()
    st.markdown("**Peso alcance:**")
    st.caption("Rango: 0 a 100 — se aplica multiplicando el porcentaje de SI en la tabla de alcance de servicios")
    _peso_alcance_pct = st.number_input(
        "Peso alcance — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, value=100, step=5,
        key="ni_peso_alcance"
    )
    peso_alcance = _peso_alcance_pct / 100

    st.markdown("**Peso metodología:**")
    st.caption("Rango: 0 a 100 — se aplica multiplicando el porcentaje de SI en la tabla de metodología implementación")
    _peso_metodologia_pct = st.number_input(
        "Peso metodología — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, value=100, step=5,
        key="ni_peso_metodologia"
    )
    peso_metodologia = _peso_metodologia_pct / 100

archivos = st.file_uploader("Sube archivos Excel", type=["xlsx"], accept_multiple_files=True)

if "archivos_cargados" not in st.session_state:
    st.session_state["archivos_cargados"] = False


# =========================
# PROCESAMIENTO
# =========================
if archivos and not st.session_state["archivos_cargados"]:
    data, data_k = [], []
    data_nf, data_k_nf = [], []
    detalles_globales, detalles_globales_k = {}, {}
    detalles_globales_nf, detalles_globales_k_nf = {}, {}
    data_experiencia = []
    data_experiencia_oferente = []
    data_alcance_servicios = []
    data_metodologia = []
    metadata_archivos = []
    nombres_proveedores = []

    for archivo in archivos:
        proveedor = Path(archivo.name).stem
        nombres_proveedores.append(proveedor)
        archivo_bytes = archivo.getvalue()
        tamano_kb = math.ceil(len(archivo_bytes) / 1024)

        fecha_modificacion = obtener_fecha_modificacion(archivo_bytes)

        metadata_archivos.append({
            "Nombre del archivo": archivo.name,
            "Tamaño (KB)": tamano_kb,
            "Fecha y hora de última modificación": fecha_modificacion,
        })

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(archivo_bytes)
            path = tmp.name

        (resultados, detalles, resultados_k, detalles_k,
         resultados_nf, detalles_nf, resultados_k_nf, detalles_k_nf) = analizar_archivo(
            path, pesos_f, peso_col_f, peso_col_g, pesos_k, incluir_calidad
        )

        for hoja, v in resultados.items():
            data.append({"Hoja": hoja, "Proveedor": proveedor, "Cumplimiento_%": v})
        for hoja, df_ in detalles.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales.setdefault(hoja, []).append(df_)
        for hoja, v in resultados_k.items():
            data_k.append({"Hoja": hoja, "Proveedor": proveedor, "Calidad_%": v})
        for hoja, df_ in detalles_k.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales_k.setdefault(hoja, []).append(df_)

        for hoja, v in resultados_nf.items():
            data_nf.append({"Hoja": hoja, "Proveedor": proveedor, "Cumplimiento_%": v})
        for hoja, df_ in detalles_nf.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales_nf.setdefault(hoja, []).append(df_)
        for hoja, v in resultados_k_nf.items():
            data_k_nf.append({"Hoja": hoja, "Proveedor": proveedor, "Calidad_%": v})
        for hoja, df_ in detalles_k_nf.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales_k_nf.setdefault(hoja, []).append(df_)

        wb_exp = openpyxl.load_workbook(path, data_only=True)
        df_exp = analizar_hoja_experiencia(wb_exp, proveedor)
        if df_exp is not None:
            data_experiencia.append(df_exp)
        df_exp_oferente = analizar_hoja_experiencia_oferente(wb_exp, proveedor)
        if df_exp_oferente is not None:
            data_experiencia_oferente.append(df_exp_oferente)
        df_alcance = analizar_hoja_alcance_servicios(wb_exp, proveedor)
        if df_alcance is not None:
            data_alcance_servicios.append(df_alcance)
        df_metodologia = analizar_hoja_metodologia(wb_exp, proveedor)
        if df_metodologia is not None:
            data_metodologia.append(df_metodologia)

    (df_final, df_total, df_final_k, df_total_k,
     df_puntaje, df_total_puntaje) = construir_tablas(
        data, data_k, peso_total_cumplimiento, peso_total_calidad, incluir_calidad
    )

    (df_final_nf, df_total_nf, df_final_k_nf, df_total_k_nf,
     df_puntaje_nf, df_total_puntaje_nf) = construir_tablas(
        data_nf, data_k_nf, peso_total_cumplimiento, peso_total_calidad, incluir_calidad
    )

    st.session_state.update({
        "df_final": df_final, "df_total": df_total,
        "df_final_k": df_final_k, "df_total_k": df_total_k,
        "df_puntaje": df_puntaje, "df_total_puntaje": df_total_puntaje,
        "df_final_nf": df_final_nf, "df_total_nf": df_total_nf,
        "df_final_k_nf": df_final_k_nf, "df_total_k_nf": df_total_k_nf,
        "df_puntaje_nf": df_puntaje_nf, "df_total_puntaje_nf": df_total_puntaje_nf,
        "detalles_globales": detalles_globales,
        "detalles_globales_k": detalles_globales_k,
        "detalles_globales_nf": detalles_globales_nf,
        "detalles_globales_k_nf": detalles_globales_k_nf,
        "data_experiencia": data_experiencia,
        "data_experiencia_oferente": data_experiencia_oferente,
        "data_alcance_servicios": data_alcance_servicios,
        "data_metodologia": data_metodologia,
        "metadata_archivos": metadata_archivos,
        "nombres_proveedores": nombres_proveedores,
        "param_incluir_calidad": incluir_calidad,
        "param_peso_col_f_raw": peso_col_f_pct,
        "param_peso_col_g_raw": peso_col_g_pct,
        "param_pesos_f_raw": {
            "SI":         _si_pct,
            "DESARROLLO": _desarrollo_pct,
            "TERCERO":    _tercero_pct,
            "NO":         _no_pct,
        },
        "param_pesos_k_raw": {
            "COMPLETO":   _k_completo   if incluir_calidad else 100,
            "PARCIAL":    _k_parcial    if incluir_calidad else 50,
            "INCOMPLETO": _k_incompleto if incluir_calidad else 0,
        },
        "param_peso_total_cumplimiento_raw": _peso_total_cumplimiento_pct if incluir_calidad else 100,
        "param_peso_total_calidad_raw":      _peso_total_calidad_pct      if incluir_calidad else 100,
        "archivos_cargados": True,
        "analisis_con_calidad": incluir_calidad,
        "param_peso_alcance_raw": _peso_alcance_pct,
        "param_peso_metodologia_raw": _peso_metodologia_pct,
    })


# =========================
# MOSTRAR
# =========================
if st.session_state["archivos_cargados"]:

    df_final            = st.session_state["df_final"]
    df_total            = st.session_state["df_total"]
    df_final_k          = st.session_state["df_final_k"]
    df_total_k          = st.session_state["df_total_k"]
    df_puntaje          = st.session_state["df_puntaje"]
    df_total_puntaje    = st.session_state["df_total_puntaje"]
    df_final_nf         = st.session_state["df_final_nf"]
    df_total_nf         = st.session_state["df_total_nf"]
    df_final_k_nf       = st.session_state["df_final_k_nf"]
    df_total_k_nf       = st.session_state["df_total_k_nf"]
    df_puntaje_nf       = st.session_state["df_puntaje_nf"]
    df_total_puntaje_nf = st.session_state["df_total_puntaje_nf"]
    detalles_globales      = st.session_state["detalles_globales"]
    detalles_globales_k    = st.session_state["detalles_globales_k"]
    detalles_globales_nf   = st.session_state["detalles_globales_nf"]
    detalles_globales_k_nf = st.session_state["detalles_globales_k_nf"]
    data_experiencia         = st.session_state["data_experiencia"]
    data_experiencia_oferente = st.session_state.get("data_experiencia_oferente", [])
    data_alcance_servicios    = st.session_state.get("data_alcance_servicios", [])
    analisis_con_calidad   = st.session_state.get("analisis_con_calidad", False)
    metadata_archivos      = st.session_state.get("metadata_archivos", [])
    nombres_proveedores    = st.session_state.get("nombres_proveedores", [])

    # ---- FUNCIONAL ----
    st.subheader("Cumplimiento funcional")

    st.markdown("#### Cumplimiento por hoja")
    event = st.dataframe(formatear_porcentaje_df(df_final), on_select="rerun", key="df_func")
    boton_descarga("⬇️ Descargar", {"Cumplimiento por hoja": df_final}, "f_cumplimiento_hoja.xlsx", "dl_f_cum_hoja")
    if event.selection.rows:
        hoja = df_final.iloc[event.selection.rows[0]]["Hoja"]
        st.session_state["detalle_hoja"] = hoja
        st.session_state["detalle_df"] = pd.concat(detalles_globales[hoja])
        st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k[hoja]) if (analisis_con_calidad and hoja in detalles_globales_k) else None
        st.switch_page("pages/detalle_hoja.py")

    st.markdown("#### Pesos por hoja funcional")
    st.caption("Rango: 0 a 100 — indica el peso porcentual de cada hoja en el total funcional")
    hojas_func_list = df_final["Hoja"].tolist()

    for hoja_w in hojas_func_list:
        if f"peso_hoja_func_{hoja_w}" not in st.session_state:
            st.session_state[f"peso_hoja_func_{hoja_w}"] = 100

    pesos_hojas_func = {}
    for hoja_w in hojas_func_list:
        col_nombre, col_input = st.columns([2, 3])
        with col_nombre:
            st.markdown(f"<div style='padding-top:8px'>{hoja_w}</div>", unsafe_allow_html=True)
        with col_input:
            pesos_hojas_func[hoja_w] = st.number_input(
                label=hoja_w,
                min_value=0,
                max_value=100,
                value=st.session_state.get(f"peso_hoja_func_{hoja_w}", 100),
                step=1,
                key=f"peso_hoja_func_{hoja_w}",
                label_visibility="collapsed"
            )

    _, col_btn_func, _ = st.columns([2, 1, 2])
    with col_btn_func:
        if st.button("Generar total cumplimiento funcional sin calidad", key="btn_total_func", use_container_width=True):
            st.session_state["mostrar_total_func"] = True
            proveedores_func = [c for c in df_final.columns if c != "Hoja"]
            df_idx_func = df_final.set_index("Hoja")
            pesos_actuales_func = {h: st.session_state.get(f"peso_hoja_func_{h}", 100) for h in hojas_func_list}
            filas_ponderadas_func = []
            for h in hojas_func_list:
                if h not in df_idx_func.index:
                    continue
                fila = {"Hoja": h}
                for prov in proveedores_func:
                    fila[prov] = round(df_idx_func.loc[h, prov] * (pesos_actuales_func[h] / 100), 2)
                filas_ponderadas_func.append(fila)
            total_ponderado_func = {"Hoja": "TOTAL"}
            for prov in proveedores_func:
                total_ponderado_func[prov] = round(
                    sum(
                        (df_idx_func.loc[h, prov] / 100) * (pesos_actuales_func[h] / 100)
                        for h in hojas_func_list if h in df_idx_func.index
                    ) * 100, 2
                )
            df_func_ponderado = pd.DataFrame(filas_ponderadas_func + [total_ponderado_func])
            st.session_state["df_total_func_ponderado"] = df_func_ponderado
            st.session_state["snapshot_pesos_hojas_func"] = dict(pesos_actuales_func)

    if st.session_state.get("mostrar_total_func", False):
        st.markdown("#### Total de cumplimiento funcional sin calidad")
        df_mostrar_total_func = st.session_state.get("df_total_func_ponderado", df_total)
        st.dataframe(formatear_porcentaje_df(df_mostrar_total_func), key="df_total_func")
        boton_descarga("⬇️ Descargar", {"Total cumplimiento": df_mostrar_total_func}, "f_total_cumplimiento.xlsx", "dl_f_total_cum")

    if analisis_con_calidad and df_final_k is not None:
        st.markdown("#### Calidad por hoja")
        event_k = st.dataframe(formatear_porcentaje_df(df_final_k), on_select="rerun", key="df_cal_func")
        boton_descarga("⬇️ Descargar", {"Calidad por hoja": df_final_k}, "f_calidad_hoja.xlsx", "dl_f_cal_hoja")
        if event_k.selection.rows:
            hoja_k = df_final_k.iloc[event_k.selection.rows[0]]["Hoja"]
            st.session_state["detalle_hoja"] = hoja_k
            st.session_state["detalle_df"] = pd.concat(detalles_globales[hoja_k]) if hoja_k in detalles_globales else None
            st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k[hoja_k])
            st.switch_page("pages/detalle_hoja.py")

        st.markdown("#### Total de calidad")
        st.dataframe(formatear_porcentaje_df(df_total_k), key="df_total_cal_func")
        boton_descarga("⬇️ Descargar", {"Total calidad": df_total_k}, "f_total_calidad.xlsx", "dl_f_total_cal")

        st.markdown("#### Puntaje funcional por hoja")
        st.dataframe(formatear_porcentaje_df(df_puntaje), use_container_width=True, key="df_punt_func")
        boton_descarga("⬇️ Descargar", {"Puntaje funcional": df_puntaje}, "f_puntaje_hoja.xlsx", "dl_f_punt_hoja")

        st.markdown("#### Total puntaje funcional")
        st.dataframe(formatear_porcentaje_df(df_total_puntaje), use_container_width=True, key="df_total_punt_func")
        boton_descarga("⬇️ Descargar", {"Total puntaje funcional": df_total_puntaje}, "f_total_puntaje.xlsx", "dl_f_total_punt")

    # ---- NO FUNCIONAL ----
    st.subheader("Cumplimiento no funcional")

    st.markdown("#### Cumplimiento por hoja")
    event_nf = st.dataframe(formatear_porcentaje_df(df_final_nf), on_select="rerun", key="df_nofunc")
    boton_descarga("⬇️ Descargar", {"Cumplimiento por hoja": df_final_nf}, "nf_cumplimiento_hoja.xlsx", "dl_nf_cum_hoja")
    if event_nf.selection.rows:
        hoja_nf = df_final_nf.iloc[event_nf.selection.rows[0]]["Hoja"]
        st.session_state["detalle_hoja"] = hoja_nf
        st.session_state["detalle_df"] = pd.concat(detalles_globales_nf[hoja_nf])
        st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k_nf[hoja_nf]) if (analisis_con_calidad and hoja_nf in detalles_globales_k_nf) else None
        st.switch_page("pages/detalle_hoja.py")

    st.markdown("#### Pesos por hoja no funcional")
    st.caption("Rango: 0 a 100 — indica el peso porcentual de cada hoja en el total no funcional")
    hojas_nofunc_list = df_final_nf["Hoja"].tolist()

    for hoja_w in hojas_nofunc_list:
        if f"peso_hoja_nf_{hoja_w}" not in st.session_state:
            st.session_state[f"peso_hoja_nf_{hoja_w}"] = 100

    pesos_hojas_nf = {}
    for hoja_w in hojas_nofunc_list:
        col_nombre, col_input = st.columns([2, 3])
        with col_nombre:
            st.markdown(f"<div style='padding-top:8px'>{hoja_w}</div>", unsafe_allow_html=True)
        with col_input:
            pesos_hojas_nf[hoja_w] = st.number_input(
                label=hoja_w,
                min_value=0,
                max_value=100,
                value=st.session_state.get(f"peso_hoja_nf_{hoja_w}", 100),
                step=1,
                key=f"peso_hoja_nf_{hoja_w}",
                label_visibility="collapsed"
            )

    _, col_btn_nf, _ = st.columns([2, 1, 2])
    with col_btn_nf:
        if st.button("Generar total cumplimiento No funcional sin calidad", key="btn_total_nf", use_container_width=True):
            st.session_state["mostrar_total_nf"] = True
            proveedores_nf = [c for c in df_final_nf.columns if c != "Hoja"]
            df_idx_nf = df_final_nf.set_index("Hoja")
            pesos_actuales_nf = {h: st.session_state.get(f"peso_hoja_nf_{h}", 100) for h in hojas_nofunc_list}
            filas_ponderadas_nf = []
            for h in hojas_nofunc_list:
                if h not in df_idx_nf.index:
                    continue
                fila = {"Hoja": h}
                for prov in proveedores_nf:
                    fila[prov] = round(df_idx_nf.loc[h, prov] * (pesos_actuales_nf[h] / 100), 2)
                filas_ponderadas_nf.append(fila)
            total_ponderado_nf = {"Hoja": "TOTAL"}
            for prov in proveedores_nf:
                total_ponderado_nf[prov] = round(
                    sum(
                        (df_idx_nf.loc[h, prov] / 100) * (pesos_actuales_nf[h] / 100)
                        for h in hojas_nofunc_list if h in df_idx_nf.index
                    ) * 100, 2
                )
            df_nf_ponderado = pd.DataFrame(filas_ponderadas_nf + [total_ponderado_nf])
            st.session_state["df_total_nf_ponderado"] = df_nf_ponderado
            st.session_state["snapshot_pesos_hojas_nf"] = dict(pesos_actuales_nf)

    if st.session_state.get("mostrar_total_nf", False):
        st.markdown("#### Total de cumplimiento no funcional sin calidad")
        df_mostrar_total_nf = st.session_state.get("df_total_nf_ponderado", df_total_nf)
        st.dataframe(formatear_porcentaje_df(df_mostrar_total_nf), key="df_total_nofunc")
        boton_descarga("⬇️ Descargar", {"Total cumplimiento": df_mostrar_total_nf}, "nf_total_cumplimiento.xlsx", "dl_nf_total_cum")

    if analisis_con_calidad and df_final_k_nf is not None:
        st.markdown("#### Calidad por hoja")
        event_k_nf = st.dataframe(formatear_porcentaje_df(df_final_k_nf), on_select="rerun", key="df_cal_nofunc")
        boton_descarga("⬇️ Descargar", {"Calidad por hoja": df_final_k_nf}, "nf_calidad_hoja.xlsx", "dl_nf_cal_hoja")
        if event_k_nf.selection.rows:
            hoja_k_nf = df_final_k_nf.iloc[event_k_nf.selection.rows[0]]["Hoja"]
            st.session_state["detalle_hoja"] = hoja_k_nf
            st.session_state["detalle_df"] = pd.concat(detalles_globales_nf[hoja_k_nf]) if hoja_k_nf in detalles_globales_nf else None
            st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k_nf[hoja_k_nf])
            st.switch_page("pages/detalle_hoja.py")

        st.markdown("#### Total de calidad")
        st.dataframe(formatear_porcentaje_df(df_total_k_nf), key="df_total_cal_nofunc")
        boton_descarga("⬇️ Descargar", {"Total calidad": df_total_k_nf}, "nf_total_calidad.xlsx", "dl_nf_total_cal")

        st.markdown("#### Puntaje no funcional por hoja")
        st.dataframe(formatear_porcentaje_df(df_puntaje_nf), use_container_width=True, key="df_punt_nofunc")
        boton_descarga("⬇️ Descargar", {"Puntaje no funcional": df_puntaje_nf}, "nf_puntaje_hoja.xlsx", "dl_nf_punt_hoja")

        st.markdown("#### Total puntaje no funcional")
        st.dataframe(formatear_porcentaje_df(df_total_puntaje_nf), use_container_width=True, key="df_total_punt_nofunc")
        boton_descarga("⬇️ Descargar", {"Total puntaje no funcional": df_total_puntaje_nf}, "nf_total_puntaje.xlsx", "dl_nf_total_punt")

    # ---- SOLIDEZ DEL FABRICANTE ----
    st.subheader("Solidez del fabricante")

    st.markdown("#### Experiencia del fabricante")
    if data_experiencia:
        df_exp_all = pd.concat(data_experiencia, ignore_index=True)
        todos_proveedores = list(df_exp_all["Proveedor"].unique())

        pivot_sector = (
            df_exp_all[df_exp_all["Sector/Industria"] != ""]
            .groupby(["Sector/Industria", "Proveedor"]).size()
            .unstack(fill_value=0)
            .reindex(columns=todos_proveedores, fill_value=0)
            .reset_index()
        )

        st.markdown("**Por Sector/Industria** — haz clic en una fila para desplegar los países")

        import json

        sectores_data = []
        for _, row_s in pivot_sector.iterrows():
            sector = row_s["Sector/Industria"]
            df_filtrado = df_exp_all[df_exp_all["Sector/Industria"] == sector]
            pivot_p = (
                df_filtrado[df_filtrado["País"] != ""]
                .groupby(["País", "Proveedor"]).size()
                .unstack(fill_value=0)
                .reindex(columns=todos_proveedores, fill_value=0)
                .reset_index()
            )
            paises = []
            for _, row_p in pivot_p.iterrows():
                paises.append({
                    "pais": row_p["País"],
                    "vals": [int(row_p[p]) for p in todos_proveedores]
                })
            sectores_data.append({
                "sector": sector,
                "vals": [int(row_s[p]) for p in todos_proveedores],
                "paises": paises
            })

        proveedores_json = json.dumps(todos_proveedores)
        sectores_json = json.dumps(sectores_data)

        tabla_html = f"""
<style>
  .wrap-exp {{
    font-family: inherit;
    border: 1px solid #3a3a4a;
    border-radius: 6px;
    overflow: hidden;
    max-height: 520px;
    overflow-y: auto;
  }}
  .exp-tbl {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13.5px;
  }}
  .exp-tbl thead tr {{
    background: #16213e;
    position: sticky;
    top: 0;
    z-index: 2;
  }}
  .exp-tbl th {{
    padding: 9px 14px;
    text-align: left;
    color: #a0aec0;
    font-weight: 600;
    border-bottom: 1px solid #3a3a4a;
    white-space: nowrap;
  }}
  .exp-tbl th.num, .exp-tbl td.num {{
    text-align: right;
  }}
  .sector-tr {{
    background: #1a1a2e;
    cursor: pointer;
    transition: background 0.15s;
  }}
  .sector-tr:hover {{
    background: #252545 !important;
  }}
  .sector-tr.open {{
    background: #1e2a45 !important;
  }}
  .sector-tr td {{
    padding: 8px 14px;
    border-bottom: 1px solid #2d2d3d;
    color: #e2e8f0;
  }}
  .arrow {{
    display: inline-block;
    width: 16px;
    font-size: 10px;
    color: #63b3ed;
    transition: transform 0.2s;
    user-select: none;
  }}
  .arrow.open {{ transform: rotate(90deg); }}
  .pais-tr {{
    display: none;
    background: #111827;
  }}
  .pais-tr.visible {{ display: table-row; }}
  .pais-tr td {{
    padding: 6px 14px 6px 38px;
    color: #90cdf4;
    font-size: 13px;
    border-bottom: 1px solid #1f2937;
  }}
  .pais-tr td.num {{ color: #7dd3fc; }}
</style>
<div class="wrap-exp">
  <table class="exp-tbl" id="expTbl">
    <thead>
      <tr>
        <th style="width:24px"></th>
        <th>Sector / Industria</th>
        {"".join(f'<th class="num">{p}</th>' for p in todos_proveedores)}
      </tr>
    </thead>
    <tbody id="expBody"></tbody>
  </table>
</div>
<script>
(function() {{
  const proveedores = {proveedores_json};
  const sectores   = {sectores_json};
  const tbody = document.getElementById("expBody");

  sectores.forEach(function(s, si) {{
    var tr = document.createElement("tr");
    tr.className = "sector-tr";
    tr.dataset.idx = si;
    var arrowId = "arr-" + si;
    var tdArrow = "<td><span class='arrow' id='" + arrowId + "'>&#9658;</span></td>";
    var tdSector = "<td>" + s.sector + "</td>";
    var tdVals = s.vals.map(function(v) {{
      return "<td class='num'>" + v + "</td>";
    }}).join("");
    tr.innerHTML = tdArrow + tdSector + tdVals;
    tr.addEventListener("click", function() {{
      var arrow = document.getElementById(arrowId);
      var isOpen = tr.classList.contains("open");
      document.querySelectorAll(".sector-tr.open").forEach(function(el) {{
        el.classList.remove("open");
      }});
      document.querySelectorAll(".arrow.open").forEach(function(el) {{
        el.classList.remove("open");
      }});
      document.querySelectorAll(".pais-tr.visible").forEach(function(el) {{
        el.classList.remove("visible");
      }});
      if (!isOpen) {{
        tr.classList.add("open");
        arrow.classList.add("open");
        document.querySelectorAll(".pais-tr[data-parent='" + si + "']").forEach(function(el) {{
          el.classList.add("visible");
        }});
      }}
    }});
    tbody.appendChild(tr);

    s.paises.forEach(function(p) {{
      var trP = document.createElement("tr");
      trP.className = "pais-tr";
      trP.dataset.parent = si;
      var tdPais = "<td></td><td>&#127758; " + p.pais + "</td>";
      var tdPVals = p.vals.map(function(v) {{
        return "<td class='num'>" + v + "</td>";
      }}).join("");
      trP.innerHTML = tdPais + tdPVals;
      tbody.appendChild(trP);
    }});
  }});
}})();
</script>
"""
        components.html(tabla_html, height=540, scrolling=False)

        boton_descarga(
            "⬇️ Descargar sectores",
            {"Experiencia por sector": pivot_sector},
            "experiencia_sector.xlsx",
            "dl_exp_sector"
        )

    else:
        st.info("No se encontraron datos de experiencia del fabricante.")

    st.markdown("#### Información de la solución - Localización Colombia/Perú")
    # tabla próximamente

    st.markdown("#### Información de la solución - Evolución")
    # tabla próximamente

    st.markdown("#### Información de la solución - Red de partners")
    # tabla próximamente

    # ---- CALIDAD DEL PROPONENTE ----
    st.subheader("Calidad del proponente")

    st.markdown("#### Experiencia del oferente")
    if data_experiencia_oferente:
        df_exp_of_all = pd.concat(data_experiencia_oferente, ignore_index=True)
        todos_proveedores_of = list(df_exp_of_all["Proveedor"].unique())

        pivot_sector_of = (
            df_exp_of_all[df_exp_of_all["Sector/Industria"] != ""]
            .groupby(["Sector/Industria", "Proveedor"]).size()
            .unstack(fill_value=0)
            .reindex(columns=todos_proveedores_of, fill_value=0)
            .reset_index()
        )

        import json as _json
        sectores_data_of = []
        for _, row_s in pivot_sector_of.iterrows():
            sector = row_s["Sector/Industria"]
            df_fil = df_exp_of_all[df_exp_of_all["Sector/Industria"] == sector]
            pivot_p = (
                df_fil[df_fil["País"] != ""]
                .groupby(["País", "Proveedor"]).size()
                .unstack(fill_value=0)
                .reindex(columns=todos_proveedores_of, fill_value=0)
                .reset_index()
            )
            paises = [
                {"pais": row_p["País"], "vals": [int(row_p[p]) for p in todos_proveedores_of]}
                for _, row_p in pivot_p.iterrows()
            ]
            sectores_data_of.append({
                "sector": sector,
                "vals": [int(row_s[p]) for p in todos_proveedores_of],
                "paises": paises
            })

        proveedores_json_of = _json.dumps(todos_proveedores_of)
        sectores_json_of    = _json.dumps(sectores_data_of)

        tabla_of_html = f"""
<style>
  .wrap-exp-of {{
    font-family: inherit;
    border: 1px solid #3a3a4a;
    border-radius: 6px;
    overflow: hidden;
    max-height: 520px;
    overflow-y: auto;
  }}
  .exp-tbl-of {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13.5px;
  }}
  .exp-tbl-of thead tr {{
    background: #16213e;
    position: sticky;
    top: 0;
    z-index: 2;
  }}
  .exp-tbl-of th {{
    padding: 9px 14px;
    text-align: left;
    color: #a0aec0;
    font-weight: 600;
    border-bottom: 1px solid #3a3a4a;
    white-space: nowrap;
  }}
  .exp-tbl-of th.num-of, .exp-tbl-of td.num-of {{ text-align: right; }}
  .sector-tr-of {{
    background: #1a1a2e;
    cursor: pointer;
    transition: background 0.15s;
  }}
  .sector-tr-of:hover {{ background: #252545 !important; }}
  .sector-tr-of.open-of {{ background: #1e2a45 !important; }}
  .sector-tr-of td {{
    padding: 8px 14px;
    border-bottom: 1px solid #2d2d3d;
    color: #e2e8f0;
  }}
  .arrow-of {{
    display: inline-block;
    width: 16px;
    font-size: 10px;
    color: #63b3ed;
    transition: transform 0.2s;
    user-select: none;
  }}
  .arrow-of.open-of {{ transform: rotate(90deg); }}
  .pais-tr-of {{ display: none; background: #111827; }}
  .pais-tr-of.visible-of {{ display: table-row; }}
  .pais-tr-of td {{
    padding: 6px 14px 6px 38px;
    color: #90cdf4;
    font-size: 13px;
    border-bottom: 1px solid #1f2937;
  }}
  .pais-tr-of td.num-of {{ color: #7dd3fc; }}
</style>
<div class="wrap-exp-of">
  <table class="exp-tbl-of" id="expTblOf">
    <thead>
      <tr>
        <th style="width:24px"></th>
        <th>Sector / Industria</th>
        {"".join(f'<th class="num-of">{p}</th>' for p in todos_proveedores_of)}
      </tr>
    </thead>
    <tbody id="expBodyOf"></tbody>
  </table>
</div>
<script>
(function() {{
  var proveedores = {proveedores_json_of};
  var sectores   = {sectores_json_of};
  var tbody = document.getElementById("expBodyOf");
  sectores.forEach(function(s, si) {{
    var tr = document.createElement("tr");
    tr.className = "sector-tr-of";
    tr.dataset.idx = si;
    var arrowId = "arr-of-" + si;
    tr.innerHTML = "<td><span class='arrow-of' id='" + arrowId + "'>&#9658;</span></td>"
      + "<td>" + s.sector + "</td>"
      + s.vals.map(function(v) {{ return "<td class='num-of'>" + v + "</td>"; }}).join("");
    tr.addEventListener("click", function() {{
      var arrow = document.getElementById(arrowId);
      var isOpen = tr.classList.contains("open-of");
      document.querySelectorAll(".sector-tr-of.open-of").forEach(function(el) {{ el.classList.remove("open-of"); }});
      document.querySelectorAll(".arrow-of.open-of").forEach(function(el) {{ el.classList.remove("open-of"); }});
      document.querySelectorAll(".pais-tr-of.visible-of").forEach(function(el) {{ el.classList.remove("visible-of"); }});
      if (!isOpen) {{
        tr.classList.add("open-of");
        arrow.classList.add("open-of");
        document.querySelectorAll(".pais-tr-of[data-parent='" + si + "']").forEach(function(el) {{ el.classList.add("visible-of"); }});
      }}
    }});
    tbody.appendChild(tr);
    s.paises.forEach(function(p) {{
      var trP = document.createElement("tr");
      trP.className = "pais-tr-of";
      trP.dataset.parent = si;
      trP.innerHTML = "<td></td><td>&#127758; " + p.pais + "</td>"
        + p.vals.map(function(v) {{ return "<td class='num-of'>" + v + "</td>"; }}).join("");
      tbody.appendChild(trP);
    }});
  }});
}})();
</script>
"""
        st.markdown("**Por Sector/Industria** — haz clic en una fila para desplegar los países")
        components.html(tabla_of_html, height=540, scrolling=False)
        boton_descarga(
            "⬇️ Descargar sectores oferente",
            {"Experiencia por sector": pivot_sector_of},
            "experiencia_sector_oferente.xlsx",
            "dl_exp_sector_of"
        )
    else:
        st.info("No se encontraron datos de experiencia del oferente (hoja '5.').")

    # ---- ALCANCE DE SERVICIOS ----
    st.markdown("#### Alcance de servicios")
    if data_alcance_servicios:
        df_alc_all = pd.concat(data_alcance_servicios, ignore_index=True)
        df_alc_all = df_alc_all[df_alc_all["Respuesta"] != ""]

        todos_provs_alc = nombres_proveedores if nombres_proveedores else sorted(df_alc_all["Proveedor"].unique())

        conteo = (
            df_alc_all.groupby(["Proveedor", "Respuesta"]).size()
            .unstack(fill_value=0)
        )
        for col in ["SI", "NO"]:
            if col not in conteo.columns:
                conteo[col] = 0
        conteo = conteo[["SI", "NO"]]
        conteo = conteo.reindex(todos_provs_alc, fill_value=0)
        total_por_prov = conteo.sum(axis=1)

        _pa = st.session_state.get("param_peso_alcance_raw", _peso_alcance_pct) / 100

        filas = []
        for resp in ["SI"]:
            fila = {"Respuesta": resp}
            for prov in todos_provs_alc:
                total = total_por_prov[prov]
                pct = round(conteo.loc[prov, resp] / total * 100, 2) if total > 0 else 0.0
                pct = round(pct * _pa, 2)
                fila[prov] = f"{pct:.2f}%"
            filas.append(fila)

        df_alcance_tabla = pd.DataFrame(filas)
        st.dataframe(df_alcance_tabla, use_container_width=True, key="df_alcance_servicios")

        filas_raw = []
        for resp in ["SI"]:
            fila = {"Respuesta": resp}
            for prov in todos_provs_alc:
                total = total_por_prov[prov]
                pct = round(conteo.loc[prov, resp] / total * 100, 2) if total > 0 else 0.0
                pct = round(pct * _pa, 2)
                fila[prov] = pct
            filas_raw.append(fila)
        df_alcance_raw = pd.DataFrame(filas_raw)
        boton_descarga(
            "⬇️ Descargar alcance de servicios",
            {"Alcance de servicios": df_alcance_raw},
            "alcance_servicios.xlsx",
            "dl_alcance_servicios"
        )
    else:
        st.info("No se encontraron datos de alcance de servicios (hoja '6.').")

    st.markdown("#### Metodología Implementación")
    data_metodologia = st.session_state.get("data_metodologia", [])
    if data_metodologia:
        df_met_all = pd.concat(data_metodologia, ignore_index=True)
        df_met_all = df_met_all[df_met_all["Respuesta"] != ""]

        todos_provs_met = nombres_proveedores if nombres_proveedores else sorted(df_met_all["Proveedor"].unique())

        conteo_met = (
            df_met_all.groupby(["Proveedor", "Respuesta"]).size()
            .unstack(fill_value=0)
        )
        for col in ["SI", "NO"]:
            if col not in conteo_met.columns:
                conteo_met[col] = 0
        conteo_met = conteo_met[["SI", "NO"]].reindex(todos_provs_met, fill_value=0)
        total_por_prov_met = conteo_met.sum(axis=1)

        _pm = st.session_state.get("param_peso_metodologia_raw", _peso_metodologia_pct) / 100

        filas_met = []
        for resp in ["SI"]:
            fila = {"Respuesta": resp}
            for prov in todos_provs_met:
                total = total_por_prov_met[prov]
                pct = round(conteo_met.loc[prov, resp] / total * 100, 2) if total > 0 else 0.0
                pct = round(pct * _pm, 2)
                fila[prov] = f"{pct:.2f}%"
            filas_met.append(fila)

        st.dataframe(pd.DataFrame(filas_met), use_container_width=True, key="df_metodologia")

        filas_met_raw = []
        for resp in ["SI"]:
            fila = {"Respuesta": resp}
            for prov in todos_provs_met:
                total = total_por_prov_met[prov]
                pct = round(conteo_met.loc[prov, resp] / total * 100, 2) if total > 0 else 0.0
                pct = round(pct * _pm, 2)
                fila[prov] = pct
            filas_met_raw.append(fila)

        boton_descarga(
            "⬇️ Descargar metodología implementación",
            {"Metodología Implementación": pd.DataFrame(filas_met_raw)},
            "metodologia_implementacion.xlsx",
            "dl_metodologia"
        )
    else:
        st.info("No se encontraron datos de metodología (hoja '7.').")

    st.markdown("#### Equipo Implementador")
    # tabla próximamente

    # ---- EXPORTAR EXCEL COMPLETO ----
    st.divider()

    pesos_hojas_func_reporte = st.session_state.get(
        "snapshot_pesos_hojas_func",
        {h: st.session_state.get(f"peso_hoja_func_{h}", 100) for h in df_final["Hoja"].tolist()}
    )
    pesos_hojas_nf_reporte = st.session_state.get(
        "snapshot_pesos_hojas_nf",
        {h: st.session_state.get(f"peso_hoja_nf_{h}", 100) for h in df_final_nf["Hoja"].tolist()}
    )

    fecha_generacion = datetime.now(tz=ZoneInfo("America/Bogota")).strftime("%Y-%m-%d %H:%M:%S")

    bloques_info = construir_hoja_info_analisis(
        fecha_generacion=fecha_generacion,
        incluir_calidad=st.session_state.get("param_incluir_calidad", analisis_con_calidad),
        peso_col_f=st.session_state.get("param_peso_col_f_raw", 100),
        peso_col_g=st.session_state.get("param_peso_col_g_raw", 100),
        pesos_f=st.session_state.get("param_pesos_f_raw", {"SI": 100, "DESARROLLO": 50, "TERCERO": 50, "NO": 0}),
        pesos_k=st.session_state.get("param_pesos_k_raw", {"COMPLETO": 100, "PARCIAL": 50, "INCOMPLETO": 0}),
        peso_total_cumplimiento=st.session_state.get("param_peso_total_cumplimiento_raw", 100),
        peso_total_calidad=st.session_state.get("param_peso_total_calidad_raw", 100),
        pesos_hojas_func=pesos_hojas_func_reporte,
        pesos_hojas_nf=pesos_hojas_nf_reporte,
        metadata_archivos=metadata_archivos,
        peso_alcance=st.session_state.get("param_peso_alcance_raw", 100),          # ← NUEVO
        peso_metodologia=st.session_state.get("param_peso_metodologia_raw", 100),  # ← NUEVO
    )

    if not analisis_con_calidad:
        df_f_total_export = st.session_state.get("df_total_func_ponderado", df_total)
        df_nf_total_export = st.session_state.get("df_total_nf_ponderado", df_total_nf)
    else:
        df_f_total_export = df_total
        df_nf_total_export = df_total_nf

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="F - Comparativo")
        df_f_total_export.to_excel(writer, index=False, sheet_name="F - Total")
        if analisis_con_calidad and df_final_k is not None:
            df_final_k.to_excel(writer, index=False, sheet_name="F - Calidad por hoja")
            df_total_k.to_excel(writer, index=False, sheet_name="F - Total calidad")
            df_puntaje.to_excel(writer, index=False, sheet_name="F - Puntaje funcional")
            df_total_puntaje.to_excel(writer, index=False, sheet_name="F - Total puntaje")
        df_final_nf.to_excel(writer, index=False, sheet_name="NF - Comparativo")
        df_nf_total_export.to_excel(writer, index=False, sheet_name="NF - Total")
        if analisis_con_calidad and df_final_k_nf is not None:
            df_final_k_nf.to_excel(writer, index=False, sheet_name="NF - Calidad por hoja")
            df_total_k_nf.to_excel(writer, index=False, sheet_name="NF - Total calidad")
            df_puntaje_nf.to_excel(writer, index=False, sheet_name="NF - Puntaje")
            df_total_puntaje_nf.to_excel(writer, index=False, sheet_name="NF - Total puntaje")
        if data_experiencia:
            pivot_sector.to_excel(writer, index=False, sheet_name="Exp - Por sector")

        if data_alcance_servicios:
            _alc_all = pd.concat(data_alcance_servicios, ignore_index=True)
            _alc_all = _alc_all[_alc_all["Respuesta"] != ""]
            _provs_alc = nombres_proveedores if nombres_proveedores else sorted(_alc_all["Proveedor"].unique())
            _conteo_alc = (
                _alc_all.groupby(["Proveedor", "Respuesta"]).size()
                .unstack(fill_value=0)
            )
            for _col in ["SI", "NO"]:
                if _col not in _conteo_alc.columns:
                    _conteo_alc[_col] = 0
            _conteo_alc = _conteo_alc[["SI", "NO"]].reindex(_provs_alc, fill_value=0)
            _total_alc = _conteo_alc.sum(axis=1)
            _filas_alc = []
            for _resp in ["SI"]:
                _fila = {"Respuesta": _resp}
                for _prov in _provs_alc:
                    _t = _total_alc[_prov]
                    _pct = round(_conteo_alc.loc[_prov, _resp] / _t * 100, 2) if _t > 0 else 0.0
                    _fila[_prov] = round(_pct * (_peso_alcance_pct / 100), 2)
                _filas_alc.append(_fila)
            pd.DataFrame(_filas_alc).to_excel(writer, index=False, sheet_name="Alcance de servicios")

        if data_metodologia:
            _met_all = pd.concat(data_metodologia, ignore_index=True)
            _met_all = _met_all[_met_all["Respuesta"] != ""]
            _provs_met = nombres_proveedores if nombres_proveedores else sorted(_met_all["Proveedor"].unique())
            _conteo_met = (
                _met_all.groupby(["Proveedor", "Respuesta"]).size()
                .unstack(fill_value=0)
            )
            for _col in ["SI", "NO"]:
                if _col not in _conteo_met.columns:
                    _conteo_met[_col] = 0
            _conteo_met = _conteo_met[["SI", "NO"]].reindex(_provs_met, fill_value=0)
            _total_met = _conteo_met.sum(axis=1)
            _filas_met = []
            for _resp in ["SI"]:
                _fila = {"Respuesta": _resp}
                for _prov in _provs_met:
                    _t = _total_met[_prov]
                    _pct = round(_conteo_met.loc[_prov, _resp] / _t * 100, 2) if _t > 0 else 0.0
                    _fila[_prov] = round(_pct * (_peso_metodologia_pct / 100), 2)
                _filas_met.append(_fila)
            pd.DataFrame(_filas_met).to_excel(writer, index=False, sheet_name="Metodologia Implementacion")

        if data_experiencia_oferente:
            _df_of_all = pd.concat(data_experiencia_oferente, ignore_index=True)
            _provs_of = list(_df_of_all["Proveedor"].unique())
            _pivot_of = (
                _df_of_all[_df_of_all["Sector/Industria"] != ""]
                .groupby(["Sector/Industria", "Proveedor"]).size()
                .unstack(fill_value=0)
                .reindex(columns=_provs_of, fill_value=0)
                .reset_index()
            )
            _pivot_of.to_excel(writer, index=False, sheet_name="Exp - Oferente por sector")

        escribir_hoja_info_analisis(writer, bloques_info)

    st.download_button(
        "⬇️ Descargar reporte completo Excel",
        buffer.getvalue(),
        file_name="reporte.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_reporte_completo"
    )